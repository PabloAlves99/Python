# pylint: disable=missing-docstring
# type: ignore
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)
# worksheet: Worksheet = workbook.active

# Nome para a planilha
SHEET_NAME = 'Minha planilha'

# Selecionou a planilha
worksheet: Worksheet = workbook[SHEET_NAME]

row: tuple[Cell]
for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Eva':
            worksheet.cell(cell.row, 2, 25)
    print()

    worksheet['B3'].value = 30

workbook.save(WORKBOOK_PATH)
