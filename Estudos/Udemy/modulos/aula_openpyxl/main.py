# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
worksheet: Worksheet = workbook.active
# Criando cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'idade')
worksheet.cell(1, 3, 'nota')

students = [
    ['Alice', 23, 8.5],
    ['Bob', 21, 7.2],
    ['Charlie', 24, 9.0],
    ['David', 22, 6.8],
    ['Eva', 20, 8.2],
    ['Frank', 25, 9.5],
    ['Grace', 19, 7.0],
    ['Hank', 23, 8.8],
    ['Ivy', 20, 7.5],
    ['Jack', 24, 9.2]
]

for i, student_row in enumerate(students, start=2):
    for j, student_column in enumerate(student_row, start=1):
        worksheet.cell(i, j, student_column)

workbook.save(WORKBOOK_PATH)
